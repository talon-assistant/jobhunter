"""XLSX export for job tracking and unemployment reporting."""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

log = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def export_applications(
    applications: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """Export applications to an XLSX workbook for unemployment reporting.

    Creates two sheets:
    1. 'Jobs Applied For' -- main tracking sheet
    2. 'Search Activity' -- summary of sources used

    Returns the output Path.
    """
    wb = Workbook()

    # --- Sheet 1: Jobs Applied For ---
    ws = wb.active
    ws.title = "Jobs Applied For"

    headers = [
        "Company", "Position", "Date Found", "Date Applied",
        "Source", "Method", "Location", "Status", "Fit Score", "Notes",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    ws.freeze_panes = "A2"

    for row_idx, app in enumerate(applications, 2):
        ws.cell(row=row_idx, column=1, value=app.get("company", ""))
        ws.cell(row=row_idx, column=2, value=app.get("position", ""))
        ws.cell(row=row_idx, column=3, value=_format_date(app.get("date_found")))
        ws.cell(row=row_idx, column=4, value=_format_date(app.get("date_applied")))
        ws.cell(row=row_idx, column=5, value=app.get("source", ""))
        ws.cell(row=row_idx, column=6, value="")  # method placeholder
        ws.cell(row=row_idx, column=7, value=app.get("location", ""))
        ws.cell(row=row_idx, column=8, value=_display_status(app.get("status", "")))
        ws.cell(row=row_idx, column=9, value=app.get("fit_score", 0))
        ws.cell(row=row_idx, column=10, value=app.get("notes", "")[:200])

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    # --- Sheet 2: Search Activity ---
    ws2 = wb.create_sheet("Search Activity")

    headers2 = ["Source", "Jobs Found", "Average Fit Score"]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Aggregate by source
    source_stats: dict[str, dict[str, Any]] = {}
    for app in applications:
        src = app.get("source", "unknown")
        if src not in source_stats:
            source_stats[src] = {"count": 0, "total_score": 0}
        source_stats[src]["count"] += 1
        source_stats[src]["total_score"] += app.get("fit_score", 0)

    for row_idx, (source, stats) in enumerate(sorted(source_stats.items()), 2):
        ws2.cell(row=row_idx, column=1, value=source)
        ws2.cell(row=row_idx, column=2, value=stats["count"])
        avg = stats["total_score"] / stats["count"] if stats["count"] else 0
        ws2.cell(row=row_idx, column=3, value=round(avg, 1))

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    log.info("Exported %d applications to %s", len(applications), out)
    return out


def _format_date(value: str | None) -> str:
    """Format an ISO date string for display."""
    if not value:
        return ""
    try:
        dt = datetime.fromisoformat(value)
        return dt.strftime("%m/%d/%Y")
    except ValueError:
        return value


def _display_status(status: str) -> str:
    """Map internal status to display string."""
    return {
        "new": "",
        "reviewing": "Reviewing",
        "applied": "Applied",
        "interviewing": "Interviewing",
        "offer": "Offer",
        "rejected": "Rejected",
        "withdrawn": "Withdrawn",
        "archived": "Archived",
    }.get(status, status)
