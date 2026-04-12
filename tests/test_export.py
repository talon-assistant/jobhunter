"""Tests for export.py."""

from jobhunter.core.export import export_applications


def test_export_xlsx(tmp_path):
    apps = [
        {
            "company": "Acme", "position": "Engineer",
            "date_found": "2026-01-15", "date_applied": "2026-01-20",
            "source": "dice", "location": "NYC",
            "status": "applied", "fit_score": 85, "notes": "Good fit",
        },
        {
            "company": "Beta Corp", "position": "Manager",
            "date_found": "2026-01-18", "date_applied": None,
            "source": "linkedin", "location": "Remote",
            "status": "new", "fit_score": 60, "notes": "",
        },
    ]
    out = export_applications(apps, tmp_path / "export.xlsx")
    assert out.exists()
    assert out.suffix == ".xlsx"

    # Verify it's a valid XLSX
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    assert "Jobs Applied For" in wb.sheetnames
    assert "Search Activity" in wb.sheetnames

    ws = wb["Jobs Applied For"]
    assert ws.cell(1, 1).value == "Company"
    assert ws.cell(2, 1).value == "Acme"
